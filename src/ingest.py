"""
Data ingestion pipeline for NUST Smart Banker.

Responsibilities:
  1. Load documents from JSON (FAQ) and XLSX (Product Knowledge) files.
  2. Anonymise PII using Microsoft Presidio before any text is stored.
  3. Chunk documents with RecursiveCharacterTextSplitter.
  4. Embed chunks with BAAI/bge-m3 via the shared Retriever.
  5. Upsert vectors + payloads into the Qdrant collection (persisted to disk).

Usage:
    python -m src.ingest               # ingest default data files
    python -m src.ingest --file path   # ingest a single new file (real-time update)
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import uuid
from pathlib import Path
from typing import List, Optional

import openpyxl
from langchain_core.documents import Document

from configs.settings import (
    DATA_DIR,
    PII_ENTITIES,
    UPLOADED_DOCS_DIR,
    XLSX_SKIP_SHEETS,
)
from src.utils import clean_text, format_qa_chunk, is_meaningful, split_documents

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")


# ─── PII Anonymisation ────────────────────────────────────────────────────────


def _build_anonymizer():
    """
    Lazily build Presidio AnalyzerEngine + AnonymizerEngine.
    Returns (analyzer, anonymizer) or (None, None) if Presidio is unavailable.
    """
    try:
        from presidio_analyzer import AnalyzerEngine
        from presidio_anonymizer import AnonymizerEngine

        analyzer = AnalyzerEngine()
        anonymizer = AnonymizerEngine()
        logger.info("Presidio anonymizer initialised.")
        return analyzer, anonymizer
    except Exception as exc:
        logger.warning("Presidio unavailable (%s) – PII anonymisation skipped.", exc)
        return None, None


_analyzer = None
_anonymizer = None


def _get_anonymizer():
    """Return (analyzer, anonymizer), lazily initializing on first call."""
    global _analyzer, _anonymizer
    if _analyzer is None:
        _analyzer, _anonymizer = _build_anonymizer()
    return _analyzer, _anonymizer


def anonymize_text(text: str) -> str:
    """
    Detect and replace PII entities in *text* with <ENTITY_TYPE> placeholders.
    Falls back to a regex-based approach if Presidio is unavailable.
    """
    if not text:
        return text

    # ── Presidio path ────────────────────────────────────────────────────────
    analyzer, anonymizer = _get_anonymizer()
    if analyzer and anonymizer:
        try:
            results = analyzer.analyze(
                text=text,
                entities=PII_ENTITIES,
                language="en",
            )
            if results:
                anonymized = anonymizer.anonymize(text=text, analyzer_results=results)
                return anonymized.text
            return text
        except Exception as exc:
            logger.debug("Presidio anonymization error: %s", exc)

    # ── Regex fallback ───────────────────────────────────────────────────────
    # Email
    text = re.sub(r"[\w.+-]+@[\w-]+\.[a-zA-Z]{2,}", "<EMAIL>", text)
    # Pakistani mobile numbers (+92-XXX-XXXXXXX or 03XX-XXXXXXX)
    text = re.sub(r"(\+92[-\s]?|0)3\d{2}[-\s]?\d{7}", "<PHONE>", text)
    # Generic phone numbers
    text = re.sub(r"\b\d{3,4}[-.\s]\d{3,4}[-.\s]\d{4}\b", "<PHONE>", text)
    # IBAN (Pakistan: PK + 22 chars)
    text = re.sub(r"\bPK\d{2}[A-Z0-9]{20}\b", "<IBAN>", text)
    # CNIC (XXXXX-XXXXXXX-X)
    text = re.sub(r"\b\d{5}-\d{7}-\d\b", "<CNIC>", text)

    return text


# ─── JSON FAQ Loader ──────────────────────────────────────────────────────────


def load_json_faq(path: Path) -> List[Document]:
    """
    Parse funds_transfer_app_features_faq.json and return a list of Documents.
    Each document corresponds to one Q&A pair; metadata carries the category.
    """
    logger.info("Loading FAQ JSON: %s", path)
    with open(path, encoding="utf-8") as fh:
        data = json.load(fh)

    documents: List[Document] = []
    categories = data.get("categories", [])

    for cat in categories:
        category_name = clean_text(cat.get("category", "General"))
        for item in cat.get("questions", []):
            question = clean_text(str(item.get("question", "")))
            answer = clean_text(str(item.get("answer", "")))

            if not is_meaningful(question) or not is_meaningful(answer):
                continue

            # Anonymise before storing
            question = anonymize_text(question)
            answer = anonymize_text(answer)

            content = format_qa_chunk(question, answer, source=category_name)
            documents.append(
                Document(
                    page_content=content,
                    metadata={
                        "source": "funds_transfer_faq",
                        "category": category_name,
                        "doc_type": "faq",
                    },
                )
            )

    logger.info("  → %d Q&A documents loaded from FAQ JSON.", len(documents))
    return documents


# ─── XLSX Product Knowledge Loader ───────────────────────────────────────────

# Cell fill colours used in the Product Knowledge workbook
_BLUE_FILL = "FF0070C0"  # headers and questions
_GRAY_FILL = "FFD8D8D8"  # cross-references to other sheets

# Pattern to identify blue cells that are genuine questions rather than
# answer sub-headers (e.g. "Little Champs Savings A/C", "Profit Payment").
_QUESTION_PATTERN = re.compile(
    r"^("
    r"\d+[\.\)]|"  # numbered: "1.", "2)"
    r"what|how|can|is|are|does|do|will|would|should|could|who|where|when|why|which|"
    r"i would like|please tell|do you offer|tell me about|is there|are there|"
    r".*\?$"  # ends with ?
    r")",
    re.IGNORECASE,
)

# Values that indicate a cell belongs to the Rate Sheet (mixed into product
# sheets).  When one of these is found the cell itself, the cell to its left,
# and the two cells above them (forming a 2x2 square) are all skipped.
_RATE_SHEET_VALUES = {
    "profit payment",
    "profit rate",
    "tenor",
    "tenure",
    "payout",
    "savings accounts",
    "term deposits",
    "term deposit",
    "change",
    "semi-annually",
    "quarterly",
    "maturity",
    "savings",
    "savings:",
    "monthly",
    "one year",
    "1 year",
    "two years",
    "three years",
    "five years",
    "one month",
    "six months",
    "two months",
    "three months",
    "29 days",
    "7 days",
    "pkr",
    "profit calculation",
    "monthly average balance",
    "profit payment frequency",
    "minimum age to qualify",
    "currency",
    "55 years",
    "–",
}


def _is_blue_cell(cell) -> bool:
    """Check if a cell has the blue fill colour used for questions/headers."""
    try:
        fg = cell.fill.fgColor
        return fg and fg.rgb and fg.rgb == _BLUE_FILL
    except Exception:
        return False


def _is_gray_cell(cell) -> bool:
    """Check if a cell has the gray fill colour used for cross-references."""
    try:
        fg = cell.fill.fgColor
        return fg and fg.rgb and fg.rgb == _GRAY_FILL
    except Exception:
        return False


def _looks_like_question(text: str) -> bool:
    """Return True if *text* looks like a question rather than a sub-header."""
    return bool(_QUESTION_PATTERN.match(text.strip()))


def _is_rate_sheet_cell(val: str) -> bool:
    """Return True if *val* looks like a Rate Sheet value rather than Q&A content."""
    val_lower = val.lower()
    if val_lower in _RATE_SHEET_VALUES:
        return True
    # Short numeric values like "0.19", "16.5", "0.1675"
    if len(val) < 10 and val.replace(".", "").replace("%", "").isdigit():
        return True
    return False


def _find_rate_sheet_squares(ws) -> set[tuple[int, str]]:
    """
    Scan the worksheet for Rate Sheet value cells and return the set of
    (row, column_letter) positions that should be skipped.

    For every cell whose text matches ``_RATE_SHEET_VALUES`` or is a
    short numeric rate value (e.g. "0.19", "16.5%"), we skip the cell
    itself, the cell to its left, and the three cells above (forming a
    2x4 block to cover merged-cell headers in the rate sheet).

    Blue non-question cells are already skipped by ``_extract_sheet_text()``,
    so this function only needs to handle numeric rate values and label cells
    that are not blue.
    """
    skip: set[tuple[int, str]] = set()
    col_order = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if not _is_rate_sheet_cell(val):
                continue

            r = cell.row
            c = cell.column_letter
            ci = col_order.index(c)

            # Always skip the rate sheet cell itself.
            skip.add((r, c))

            # For columns D+ (ci > 2): also skip the left cell and three
            # cells above to cover merged-cell headers in the rate sheet.
            # For columns A/B/C: only skip the cell itself — never skip
            # column B via the "left" rule to protect legitimate Q&A content.
            if ci > 2:
                skip.add((r, col_order[ci - 1]))
                for above in range(1, 4):
                    if r > above:
                        skip.add((r - above, c))
                        skip.add((r - above, col_order[ci - 1]))

    return skip


def _extract_sheet_text(ws, sheet_name: str) -> List[tuple[str, str]]:
    """
    Extract Q&A pairs from a single product worksheet using cell colours.

    Strategy:
      - Blue cells (FF0070C0) that match ``_QUESTION_PATTERN`` are questions.
        Other blue cells are answer sub-headers and are treated as content.
      - Non-blue, non-gray cells are answer content (may span multiple rows
        and multiple columns).
      - Gray cells (FFD8D8D8) are cross-references to other sheets – skip.
      - Formula references (values starting with ``='``) are skipped.
      - Rate Sheet value cells (and the 2x2 block around them) are skipped.
      - The very first blue cell is usually a product-name header; it is
        kept only if answer content follows it.
    """
    pairs: List[tuple[str, str]] = []
    current_q: Optional[str] = None
    answer_parts: List[str] = []

    rate_skip = _find_rate_sheet_squares(ws)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            # Skip rate sheet 2x2 blocks
            if (cell.row, cell.column_letter) in rate_skip:
                continue

            # Skip gray cross-reference cells and formula refs
            if _is_gray_cell(cell):
                continue
            val = str(cell.value).strip()
            if not val or val.startswith("='"):
                continue

            if _is_blue_cell(cell) and _looks_like_question(val):
                # Save the previous Q&A pair before starting a new question
                if current_q and answer_parts:
                    pairs.append(
                        (clean_text(current_q), clean_text("\n".join(answer_parts)))
                    )
                current_q = val
                answer_parts = []
            elif _is_blue_cell(cell):
                # Blue cell that is NOT a question — treat as a sub-header
                # or rate-sheet artefact; skip entirely.
                pass
            else:
                # Accumulate answer lines under the current question
                # (includes non-blue cells AND blue sub-headers)
                if current_q is not None:
                    answer_parts.append(val)

    # Flush last pair
    if current_q and answer_parts:
        pairs.append((clean_text(current_q), clean_text("\n".join(answer_parts))))

    # First pair is often just a product-name header with no real answer;
    # drop it if the "answer" is a single short fragment.
    if pairs and not is_meaningful(pairs[0][1], min_chars=20):
        pairs = pairs[1:]

    return pairs


def _extract_main_sheet(ws) -> List[tuple[str, str]]:
    """
    Convert the Main sheet (table of contents) into broad category Q&A.

    Layout (row 2 onward):
      Col B = category header or section number (blue), Col C = product name (gray)
      Col E = category header or section number (blue), Col F = product name (gray)
    """
    categories: dict[str, List[str]] = {}
    current_category_b: Optional[str] = None
    current_category_e: Optional[str] = None

    for row in ws.iter_rows(min_row=2):
        row_cells = list(row)
        if len(row_cells) < 6:
            continue

        col_b = row_cells[1]
        col_c = row_cells[2]
        col_e = row_cells[4]
        col_f = row_cells[5]

        if _is_blue_cell(col_b) and col_b.value is not None:
            cat_text = clean_text(str(col_b.value))
            if not re.match(r"^\d+\.0?$", cat_text) and not re.match(
                r"^\d+$", cat_text
            ):
                current_category_b = cat_text
                categories.setdefault(cat_text, [])

        if _is_blue_cell(col_e) and col_e.value is not None:
            cat_text = clean_text(str(col_e.value))
            if not re.match(r"^\d+\.0?$", cat_text) and not re.match(
                r"^\d+$", cat_text
            ):
                current_category_e = cat_text
                categories.setdefault(cat_text, [])

        if _is_gray_cell(col_c) and col_c.value is not None and current_category_b:
            product_name = clean_text(str(col_c.value))
            if product_name:
                categories[current_category_b].append(product_name)

        if _is_gray_cell(col_f) and col_f.value is not None and current_category_e:
            product_name = clean_text(str(col_f.value))
            if product_name:
                categories[current_category_e].append(product_name)

    pairs: List[tuple[str, str]] = []
    for category, products in categories.items():
        if not products:
            continue
        question = f"What {category.lower()} does NUST Bank offer?"
        answer = ", ".join(products)
        pairs.append((question, answer))

    return pairs


def _split_combined_product_names(product_name: str) -> list[str]:
    """
    Split product names that contain " and " into separate products.
    E.g., "NUST Bachat Account -Individual/Corporate and Value Plus Term Deposit*"
    becomes ["NUST Bachat Account -Individual/Corporate", "Value Plus Term Deposit*"]
    """
    if " and " not in product_name.lower():
        return [product_name]
    # Split on " and " (case-insensitive)
    parts = re.split(r"\s+and\s+", product_name, flags=re.IGNORECASE)
    return [part.strip() for part in parts if part.strip()]


def _extract_rate_sheet(ws) -> List[tuple[str, str]]:
    """
    Convert the Rate Sheet tabular data into Q&A pairs.

    Strategy:
      - Savings: Find "Profit Payment" in column B → previous row has product name.
        Next row(s) have payment frequency and rate.
      - Term Deposits: Find product name in column F OR detect "Term Deposits" as generic product.
        Look for "Tenor" as header, then tenor/rate pairs below.
        Products with " and " are split into separate entries.
        Multiple "Term Deposits" generic entries are numbered to avoid deduplication.
    """
    pairs: List[tuple[str, str]] = []

    current_savings_product: Optional[str] = None
    savings_payment: Optional[str] = None
    savings_rate: Optional[str] = None

    current_td_product: Optional[str] = None
    current_td_products: list[str] = []  # Handle products split by " and "
    td_entries: List[str] = []
    term_deposits_count = 0  # Track occurrences of generic "Term Deposits"

    def flush_savings():
        nonlocal current_savings_product, savings_payment, savings_rate
        if not current_savings_product or not savings_rate:
            return
        rate_pct = (
            f"{float(savings_rate) * 100:.1f}%"
            if savings_rate.replace(".", "").isdigit()
            else savings_rate
        )
        payment = f" paid {savings_payment}" if savings_payment else ""
        pairs.append(
            (
                f"What is the profit rate on {current_savings_product}?",
                f"{rate_pct}{payment}",
            )
        )

    def flush_td():
        nonlocal current_td_product, current_td_products, td_entries
        if not current_td_products or not td_entries:
            return
        # If we have multiple products (split by " and "), create entries for each
        for product in current_td_products:
            pairs.append(
                (
                    f"What are the term deposit rates for {product}?",
                    " | ".join(td_entries),
                )
            )

    rows = list(ws.iter_rows(min_row=10))
    row_indices = list(range(10, 10 + len(rows)))

    # Tenor pattern - matches various time periods (tenor values in the rate sheet)
    tenor_pattern = r"^(one|two|three|six|five|seven|7 days|seven days|29 days|29days|monthly|quarterly|semi-annually|annually|maturity|one month|two months|three months|six months|one year|two years|three years|five years)$"

    for i, row in enumerate(rows):
        row_idx = row_indices[i]
        cells = {cell.column_letter: cell for cell in row if cell.value is not None}

        b = cells.get("B")
        d = cells.get("D")
        f = cells.get("F")
        g = cells.get("G")
        i_cell = cells.get("I")

        b_val = clean_text(str(b.value)) if b else ""
        d_val = clean_text(str(d.value)) if d else ""
        f_val = clean_text(str(f.value)) if f else ""
        g_val = clean_text(str(g.value)) if g else ""
        i_val = clean_text(str(i_cell.value)) if i_cell else ""

        # Handle savings markers BEFORE skipping for TD headers
        # (some rows have both Profit Payment in B and Term Deposits in F)
        if b_val.upper() == "SAVINGS ACCOUNTS":
            continue

        # === SAVINGS SIDE: Check for "Profit Payment" marker FIRST ===
        if b_val.upper() == "PROFIT PAYMENT":
            if i > 0:
                prev_row = rows[i - 1]
                prev_cells = {
                    cell.column_letter: cell
                    for cell in prev_row
                    if cell.value is not None
                }
                prev_b = prev_cells.get("B")
                if prev_b:
                    prev_b_val = clean_text(str(prev_b.value))
                    if (
                        prev_b_val
                        and prev_b_val.upper()
                        not in ("SAVINGS ACCOUNTS", "PROFIT PAYMENT")
                        and "/" not in prev_b_val
                    ):
                        flush_savings()
                        current_savings_product = prev_b_val

        # After "Profit Payment" row, look for payment + rate
        if current_savings_product:
            if b_val.lower() in ("semi-annually", "monthly", "quarterly", "annually"):
                savings_payment = b_val
            if d_val and d_val.replace(".", "").replace("%", "").isdigit():
                savings_rate = d_val

        # === TERM DEPOSIT SIDE ===
        # Check for "Term Deposits" header - treat as product if followed by "Tenor"
        # Each occurrence is a separate product (for different savings products)
        if f_val.upper() == "TERM DEPOSITS":
            if i + 1 < len(rows):
                next_row = rows[i + 1]
                next_cells = {
                    cell.column_letter: cell
                    for cell in next_row
                    if cell.value is not None
                }
                next_f = next_cells.get("F")
                next_f_val = clean_text(str(next_f.value)) if next_f else ""
                if next_f_val.upper() == "TENOR":
                    flush_td()
                    term_deposits_count += 1
                    # Number them if there are multiple occurrences
                    product_name = (
                        "Term Deposits"
                        if term_deposits_count == 1
                        else f"Term Deposits ({term_deposits_count})"
                    )
                    current_td_product = product_name
                    current_td_products = [product_name]
                    td_entries = []
            continue

        if f_val.upper() == "TENOR":
            if i > 0:
                prev_row = rows[i - 1]
                prev_cells = {
                    cell.column_letter: cell
                    for cell in prev_row
                    if cell.value is not None
                }
                prev_f = prev_cells.get("F")
                if prev_f:
                    prev_f_val = clean_text(str(prev_f.value))
                    if (
                        prev_f_val
                        and prev_f_val.upper()
                        not in (
                            "TENOR",
                            "SHORT NOTICE DEPOSIT",
                            "TERM DEPOSITS",
                        )
                        and not prev_f_val.startswith("*")
                    ):
                        flush_td()
                        # Split products if they contain " and "
                        current_td_product = prev_f_val
                        current_td_products = _split_combined_product_names(prev_f_val)
                        td_entries = []
            continue

        # Detect TD products that appear WITHOUT a "Tenor" marker
        if f_val and not i_val:
            if (
                i > 0
                and f_val.upper()
                not in (
                    "TENOR",
                    "PROFIT RATE",
                    "PAYOUT",
                    "SHORT NOTICE DEPOSIT",
                )
                and not f_val.startswith("*")
                and f_val.upper() != "FCY"
                and f_val.upper() != "SAVINGS ACCOUNT"
            ):
                prev_row = rows[i - 1]
                prev_cells = {
                    cell.column_letter: cell
                    for cell in prev_row
                    if cell.value is not None
                }
                prev_f = prev_cells.get("F")
                prev_f_val = clean_text(str(prev_f.value)) if prev_f else ""

                if not prev_f_val or prev_f_val.upper() in (
                    "TERM DEPOSITS",
                    "",
                ):
                    if not re.match(tenor_pattern, f_val, re.IGNORECASE):
                        flush_td()
                        current_td_product = f_val
                        current_td_products = _split_combined_product_names(f_val)
                        td_entries = []

        # Collect tenor entries (column F has tenor, column I or G has rate)
        if current_td_products and f_val:
            rate_val = i_val if i_val else g_val
            if rate_val and rate_val.replace(".", "").replace("%", "").isdigit():
                if re.match(tenor_pattern, f_val, re.IGNORECASE):
                    rate_pct = f"{float(rate_val) * 100:.1f}%"
                    payout = ""
                    if g_val and g_val.upper() not in (
                        "PAYOUT",
                        "MATURITY",
                        "TENOR",
                        "PROFIT RATE",
                    ):
                        payout = f" ({g_val})"
                    td_entries.append(f"{f_val}: {rate_pct}{payout}")

    flush_savings()
    flush_td()
    return pairs


def load_xlsx_products(path: Path) -> List[Document]:
    """
    Parse NUST Bank-Product-Knowledge.xlsx.

    Each sheet is routed to the appropriate extractor:
      - Main sheet → broad category Q&A
      - Rate Sheet → profit/term-deposit rate Q&A
      - All other sheets → colour-based Q&A extraction (blue = question)

    Returns a list of Documents, one per Q&A pair extracted.
    """
    logger.info("Loading Product XLSX: %s", path)
    wb = openpyxl.load_workbook(path, data_only=True)
    documents: List[Document] = []
    seen_pairs: set[tuple[str, str]] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in XLSX_SKIP_SHEETS:
            logger.debug("  Skipping sheet: %s", sheet_name)
            continue

        # Skip the empty template sheet
        if sheet_name.strip() == "Sheet1":
            logger.debug("  Skipping empty template sheet: Sheet1")
            continue

        ws = wb[sheet_name]

        # Route to the appropriate extractor
        if sheet_name.strip() == "Main":
            pairs = _extract_main_sheet(ws)
        elif "Rate Sheet" in sheet_name:
            pairs = _extract_rate_sheet(ws)
        else:
            pairs = _extract_sheet_text(ws, sheet_name)

        for question, answer in pairs:
            # Deduplicate identical Q&A pairs across sheets
            pair_key = (question.lower(), answer.lower())
            if pair_key in seen_pairs:
                continue
            seen_pairs.add(pair_key)

            question = anonymize_text(question)
            answer = anonymize_text(answer)

            if not is_meaningful(answer):
                continue

            content = format_qa_chunk(question, answer, source=sheet_name.strip())
            documents.append(
                Document(
                    page_content=content,
                    metadata={
                        "source": "product_knowledge",
                        "product": sheet_name.strip(),
                        "doc_type": "product_faq",
                    },
                )
            )

        logger.info("  → Sheet %-25s : %d Q&A pairs", sheet_name, len(pairs))

    logger.info("  → %d total documents loaded from XLSX.", len(documents))
    return documents


# ─── Generic Plain-Text / JSON Loader (for real-time uploads) ─────────────────


def load_text_file(path: Path, source_label: str = "") -> List[Document]:
    """
    Load a plain .txt file and return it as a single Document.
    Used for real-time document uploads via the UI.
    """
    text = path.read_text(encoding="utf-8", errors="replace")
    text = clean_text(anonymize_text(text))
    if not is_meaningful(text, min_chars=50):
        return []
    label = source_label or path.stem
    return [
        Document(
            page_content=text,
            metadata={"source": label, "doc_type": "upload", "filename": path.name},
        )
    ]


def load_uploaded_json(path: Path, source_label: str = "") -> List[Document]:
    """
    Load an uploaded JSON file that follows the same FAQ schema as the main
    dataset (categories → questions), or falls back to loading it as plain text.
    """
    try:
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        if "categories" in data:
            # Reuse the FAQ loader with the new path
            return load_json_faq(path)
    except Exception:
        pass
    # Fallback: treat it as raw text
    return load_text_file(path, source_label)


# ─── Main Ingestion Entry Point ───────────────────────────────────────────────


_SUPPORTED_EXTENSIONS = {".json", ".xlsx", ".xls", ".txt"}


def ingest_all() -> int:
    """
    Auto-discover all supported files in DATA_DIR, load them, chunk,
    embed, and upsert to Qdrant. Returns the total number of chunks indexed.
    """
    from src.retriever import get_retriever

    logger.info("=== Starting full data ingestion ===")

    if not DATA_DIR.exists():
        logger.warning("DATA_DIR does not exist: %s", DATA_DIR)
        return 0

    files = sorted(
        [
            p
            for p in DATA_DIR.iterdir()
            if p.is_file() and p.suffix.lower() in _SUPPORTED_EXTENSIONS
        ]
    )

    if not files:
        logger.warning("No supported files found in %s.", DATA_DIR)
        return 0

    docs: List[Document] = []
    for path in files:
        ext = path.suffix.lower()
        try:
            if ext == ".json":
                docs.extend(load_uploaded_json(path, source_label=path.stem))
            elif ext in {".xlsx", ".xls"}:
                docs.extend(load_xlsx_products(path))
            else:
                docs.extend(load_text_file(path, source_label=path.stem))
            logger.info("  Loaded: %s", path.name)
        except Exception as exc:
            logger.warning("  Failed to load %s: %s", path.name, exc)

    if not docs:
        logger.error("No documents loaded — check data files.")
        return 0

    logger.info("Total raw documents: %d", len(docs))
    chunks = split_documents(docs)
    logger.info("Total chunks after splitting: %d", len(chunks))

    retriever = get_retriever()
    retriever.upsert_chunks(chunks)
    logger.info("=== Ingestion complete: %d chunks indexed. ===", len(chunks))
    return len(chunks)


def ingest_file(file_path: str | Path, source_label: str = "") -> int:
    """
    Ingest a single new file into the existing Qdrant collection.
    Supports .txt, .json, and .xlsx.  Used by the UI for real-time updates.
    Returns the number of newly indexed chunks.
    """
    from src.retriever import get_retriever

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    logger.info("Ingesting new file: %s", path.name)

    ext = path.suffix.lower()
    if ext == ".json":
        docs = load_uploaded_json(path, source_label or path.stem)
    elif ext in {".xlsx", ".xls"}:
        docs = load_xlsx_products(path)
    else:
        docs = load_text_file(path, source_label or path.stem)

    if not docs:
        logger.warning("No meaningful content found in %s", path.name)
        return 0

    chunks = split_documents(docs)
    retriever = get_retriever()
    retriever.upsert_chunks(chunks)
    logger.info("Ingested %d chunks from %s.", len(chunks), path.name)
    return len(chunks)


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NUST Smart Banker – Ingest data")
    parser.add_argument(
        "--file",
        type=str,
        default=None,
        help="Path to a single file to ingest (optional; omit to ingest all defaults)",
    )
    parser.add_argument(
        "--label",
        type=str,
        default="",
        help="Human-readable source label for the uploaded file",
    )
    args = parser.parse_args()

    if args.file:
        n = ingest_file(args.file, args.label)
        print(f"Indexed {n} chunks from {args.file}.")
    else:
        n = ingest_all()
        print(f"Full ingestion complete: {n} chunks indexed.")
